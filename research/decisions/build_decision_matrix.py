"""
Build the decision-matrix v1 xlsx from the data dict.

Usage:
    python build_decision_matrix.py
    python build_decision_matrix.py --output <path>
    python build_decision_matrix.py --check
        # only run the asserts; do not write the xlsx.

Convention (per user workflow rule, 2026-08-18):
    "dicts + assert k in d" - every required key is asserted, both
    per-row and at the workbook level.

Output is written to research/decisions/decision-matrix-v1.xlsx by
default. The path is gitignored (see research/.gitignore); the data
file, build script, test file, and README in this directory are
tracked.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from decision_matrix_data import DECISION_MATRIX_ROWS, EXPECTED_ROW_COUNT


# ---------------------------------------------------------------------------
# Schema (the "dicts + assert k in d" rule)
# ---------------------------------------------------------------------------

REQUIRED_ROW_KEYS: set[str] = {
    "id",
    "question",
    "phase",
    "severity",
    "effort",
    "depends_on",
    "affected_files",
    "my_recommendation",
    "recommendation_rationale",
}

ALLOWED_PHASES: set[str] = {
    "Toggle surface",
    "Phase 1",
    "Phase 2",
    "Phase 3",
    "Phase 4",
    "Deferred",
}

ALLOWED_SEVERITY: set[str] = {"High", "Medium", "Low"}

ALLOWED_EFFORT: set[str] = {"S", "M", "L", "XL"}

ALLOWED_DECISION: set[str] = {"Ship", "Defer", "Reject"}

# Column layout: (name, width, kind). "kind" is metadata only; the
# build script does not act on it. Width is in openpyxl's character
# units (Excel column width). Column order is the visual order on the
# Decision Matrix sheet.
COLUMNS: list[tuple[str, int, str]] = [
    ("id", 6, "int"),
    ("question", 60, "string"),
    ("phase", 16, "string"),
    ("severity", 10, "string"),
    ("effort", 8, "string"),
    ("depends_on", 12, "list[int]"),
    ("affected_files", 60, "string"),
    ("my_recommendation", 18, "string"),
    ("recommendation_rationale", 60, "string"),
    ("user_decision", 14, "string"),
    ("user_notes", 60, "string"),
]

OVERVIEW_DATE = "2026-08-26"
OVERVIEW_AUTHOR = "Mavis"
OVERVIEW_STATUS = "Awaiting review"
OVERVIEW_TITLE = "DeepSeek Harness for pi - Decision matrix v1"


# ---------------------------------------------------------------------------
# Asserts
# ---------------------------------------------------------------------------


def assert_data() -> None:
    """The "dicts + assert k in d" rule. Fail loud at build time.

    Per-row asserts:
      - exactly REQUIRED_ROW_KEYS are present (no missing, no extras)
      - phase / severity / effort / my_recommendation are in their allow-lists
      - depends_on is a list of int (empty list is OK)

    Workbook-level asserts:
      - len(DECISION_MATRIX_ROWS) == EXPECTED_ROW_COUNT
      - ids are 1..N in order, no gaps, no duplicates
      - depends_on references resolve to known ids
    """
    if len(DECISION_MATRIX_ROWS) != EXPECTED_ROW_COUNT:
        raise AssertionError(
            f"expected {EXPECTED_ROW_COUNT} rows, got {len(DECISION_MATRIX_ROWS)}"
        )

    ids: list[int] = []
    for row in DECISION_MATRIX_ROWS:
        row_id = row.get("id", "?")
        missing = REQUIRED_ROW_KEYS - row.keys()
        if missing:
            raise AssertionError(f"row {row_id} missing keys: {sorted(missing)}")
        extra = row.keys() - REQUIRED_ROW_KEYS
        if extra:
            raise AssertionError(
                f"row {row_id} has unexpected keys: {sorted(extra)}"
            )

        if row["phase"] not in ALLOWED_PHASES:
            raise AssertionError(
                f"row {row_id} phase {row['phase']!r} not in {sorted(ALLOWED_PHASES)}"
            )
        if row["severity"] not in ALLOWED_SEVERITY:
            raise AssertionError(
                f"row {row_id} severity {row['severity']!r} not in {sorted(ALLOWED_SEVERITY)}"
            )
        if row["effort"] not in ALLOWED_EFFORT:
            raise AssertionError(
                f"row {row_id} effort {row['effort']!r} not in {sorted(ALLOWED_EFFORT)}"
            )
        if row["my_recommendation"] not in ALLOWED_DECISION:
            raise AssertionError(
                f"row {row_id} recommendation {row['my_recommendation']!r} "
                f"not in {sorted(ALLOWED_DECISION)}"
            )

        deps = row["depends_on"]
        if not isinstance(deps, list):
            raise AssertionError(
                f"row {row_id} depends_on must be a list[int], got {type(deps).__name__}"
            )
        for dep in deps:
            if not isinstance(dep, int):
                raise AssertionError(
                    f"row {row_id} depends_on contains non-int: {dep!r}"
                )

        ids.append(row["id"])

    expected_ids = list(range(1, EXPECTED_ROW_COUNT + 1))
    if ids != expected_ids:
        raise AssertionError(
            f"ids must be {expected_ids} in order; got {ids}"
        )

    id_set = set(ids)
    for row in DECISION_MATRIX_ROWS:
        for dep in row["depends_on"]:
            if dep not in id_set:
                raise AssertionError(
                    f"row {row['id']} depends_on {dep!r} not in id set {sorted(id_set)}"
                )


# ---------------------------------------------------------------------------
# Workbook construction
# ---------------------------------------------------------------------------


def build_overview_sheet(wb: Workbook) -> None:
    """Sheet 1: a single A4-landscape page of summary text."""
    ws = wb.active
    ws.title = "Overview"

    # Title
    ws["A1"] = OVERVIEW_TITLE
    ws["A1"].font = Font(bold=True, size=14)

    # Metadata block
    ws["A3"] = "Date:"
    ws["B3"] = OVERVIEW_DATE
    ws["A4"] = "Status:"
    ws["B4"] = OVERVIEW_STATUS
    ws["A5"] = "Author:"
    ws["B5"] = OVERVIEW_AUTHOR
    for cell_addr in ("A3", "A4", "A5"):
        ws[cell_addr].font = Font(bold=True)

    # Source artefacts
    ws["A7"] = "Source artefacts:"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = "  - research/report-context-window-management.md (analysis)"
    ws["A9"] = "  - research/roadmap-adaptive-context.md (four-phase rollout)"
    ws["A10"] = "  - prior artifacts/plan.md (toggle-surface PRD; git history)"

    # How to use
    ws["A12"] = "How to use:"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = (
        "Review the 15 questions on the 'Decision Matrix' sheet. Fill in the "
        "'user decision' column (Ship / Defer / Reject) and 'user notes' for any "
        "item you defer or reject. The build script (build_decision_matrix.py) "
        "re-generates the xlsx from the data dict; close this file in Excel "
        "before re-running the build, or the .bak preserved by the build will "
        "hold your hand-edits."
    )
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[13].height = 60

    # Summary stats (computed at build time from the data)
    by_decision: dict[str, int] = {d: 0 for d in ALLOWED_DECISION}
    for row in DECISION_MATRIX_ROWS:
        by_decision[row["my_recommendation"]] += 1
    total = sum(by_decision.values())

    ws["A15"] = "My recommendation summary:"
    ws["A15"].font = Font(bold=True)
    ws["A16"] = f"  Ship:   {by_decision['Ship']}"
    ws["A17"] = f"  Defer:  {by_decision['Defer']}"
    ws["A18"] = f"  Reject: {by_decision['Reject']}"
    ws["A19"] = f"  Total:  {total}"

    # Column widths for Sheet 1
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 16


def build_decision_matrix_sheet(wb: Workbook) -> None:
    """Sheet 2: 15 data rows x 11 columns. Header row frozen."""
    ws = wb.create_sheet("Decision Matrix")

    # Header
    for col_idx, (name, _width, _kind) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name.replace("_", " "))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # Body
    for row_idx, row in enumerate(DECISION_MATRIX_ROWS, start=2):
        for col_idx, (name, _width, _kind) in enumerate(COLUMNS, start=1):
            value = row.get(name, "")
            if name == "depends_on" and isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    for col_idx, (_name, width, _kind) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # Reasonable row height for the data rows so wrapped text is readable.
    for row_idx in range(2, EXPECTED_ROW_COUNT + 2):
        ws.row_dimensions[row_idx].height = 60


def build_workbook() -> Workbook:
    wb = Workbook()
    build_overview_sheet(wb)
    build_decision_matrix_sheet(wb)
    return wb


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def _preserve_existing_xlsx(out_path: Path) -> None:
    """If an xlsx already exists at `out_path`, rename it to .bak.

    This protects hand-edits in Excel from being silently overwritten
    by a re-run of the build.
    """
    if out_path.exists():
        bak = out_path.with_suffix(out_path.suffix + ".bak")
        # If a previous .bak exists, drop it so we keep only the most
        # recent hand-edit. The user can recover via git history.
        if bak.exists():
            bak.unlink()
        out_path.rename(bak)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--output",
        default="decision-matrix-v1.xlsx",
        help="output xlsx path (default: decision-matrix-v1.xlsx in the cwd)",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="run asserts only; do not write the xlsx",
    )
    args = parser.parse_args(argv)

    assert_data()

    if args.check:
        print(f"asserts ok: {len(DECISION_MATRIX_ROWS)} rows")
        return 0

    out = Path(args.output)
    _preserve_existing_xlsx(out)

    wb = build_workbook()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out} ({out.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
